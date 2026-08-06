from datetime import date
from django.db.models import Sum, F, FloatField, Avg
import json
from django.shortcuts import render
from django.utils import timezone
from sales.models import SaleItem
from expenses.models import Expense, ExpenseCategory
from products.models import Product
from inventory.models import InventoryEntry
from .models import MonthlySummary
from django.http import HttpResponse
import csv
from django.shortcuts import get_object_or_404
from io import BytesIO


def format_cop(value):
    return f"${value:,.0f} COP".replace(',', '.')


def get_month_range(year, month):
    return date(year, month, 1)


def home(request):
    today = timezone.localdate()
    first_day_month = today.replace(day=1)
    # Spanish month names for display
    month_names_es = ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio', 'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre']
    current_month_display = f"{month_names_es[today.month - 1].capitalize()} {today.year}"
    annual_start = today.replace(month=1, day=1)

    sales_items = SaleItem.objects.filter(sale__sale_date__gte=annual_start)
    expenses_year = Expense.objects.filter(date__gte=annual_start)

    monthly_sales = sales_items.filter(sale__sale_date__gte=first_day_month)
    monthly_expenses = expenses_year.filter(date__gte=first_day_month)

    sales_total_month = monthly_sales.aggregate(total=Sum(F('unit_price') * F('quantity'), output_field=FloatField()))['total'] or 0
    expenses_total_month = monthly_expenses.aggregate(total=Sum('amount'))['total'] or 0
    sales_total_year = sales_items.aggregate(total=Sum(F('unit_price') * F('quantity'), output_field=FloatField()))['total'] or 0
    expenses_total_year = expenses_year.aggregate(total=Sum('amount'))['total'] or 0

    profit_month = sales_total_month - expenses_total_month
    profit_year = sales_total_year - expenses_total_year
    margin_month = (profit_month / sales_total_month * 100) if sales_total_month else 0
    margin_year = (profit_year / sales_total_year * 100) if sales_total_year else 0

    kilograms_sold_month = monthly_sales.aggregate(total_grams=Sum(F('product__weight_grams') * F('quantity'), output_field=FloatField()))['total_grams'] or 0
    kilograms_sold_year = sales_items.aggregate(total_grams=Sum(F('product__weight_grams') * F('quantity'), output_field=FloatField()))['total_grams'] or 0

    sales_count_month = monthly_sales.values('sale').distinct().count()
    average_ticket = (sales_total_month / sales_count_month) if sales_count_month else 0

    best_selling_product = Product.objects.annotate(total_quantity=Sum('saleitem__quantity')).order_by('-total_quantity').first()
    most_profitable_product = Product.objects.annotate(total_profit=Sum((F('saleitem__unit_price') - F('production_cost')) * F('saleitem__quantity'), output_field=FloatField())).order_by('-total_profit').first()

    sales_history = []
    expenses_history = []
    kg_history = []
    labels = []
    current_year = today.year
    current_month = today.month
    for idx in range(11, -1, -1):
        month = current_month - idx
        year = current_year
        if month <= 0:
            month += 12
            year -= 1
        # Use Spanish abbreviated month names for chart labels
        labels.append(month_names_es[month - 1].capitalize())
        period_sales = sales_items.filter(sale__sale_date__year=year, sale__sale_date__month=month)
        period_expenses = expenses_year.filter(date__year=year, date__month=month)
        sales_history.append(period_sales.aggregate(total=Sum(F('unit_price') * F('quantity'), output_field=FloatField()))['total'] or 0)
        expenses_history.append(period_expenses.aggregate(total=Sum('amount'))['total'] or 0)
        kg_history.append((period_sales.aggregate(total_grams=Sum(F('product__weight_grams') * F('quantity'), output_field=FloatField()))['total_grams'] or 0) / 1000)

    categories = ExpenseCategory.objects.all()
    expense_distribution = []
    for category in categories:
        total = expenses_year.filter(category=category).aggregate(total=Sum('amount'))['total'] or 0
        expense_distribution.append({'name': category.name, 'total': total})

    inventory_totals = InventoryEntry.objects.aggregate(
        total_bags=Sum('bags_added'),
        total_kilos=Sum('kilos_added'),
    )
    inventory_month = InventoryEntry.objects.filter(date__gte=first_day_month).aggregate(
        month_bags=Sum('bags_added'),
        month_kilos=Sum('kilos_added'),
    )
    sold_totals = SaleItem.objects.aggregate(
        sold_bags=Sum('quantity'),
        sold_kilos=Sum(F('quantity') * F('product__weight_grams') / 1000.0, output_field=FloatField()),
    )
    sold_monthly = SaleItem.objects.filter(sale__sale_date__gte=first_day_month).aggregate(
        sold_bags=Sum('quantity'),
        sold_kilos=Sum(F('quantity') * F('product__weight_grams') / 1000.0, output_field=FloatField()),
    )

    total_bags = inventory_totals['total_bags'] or 0
    total_kilos = inventory_totals['total_kilos'] or 0
    sold_bags_total = sold_totals['sold_bags'] or 0
    sold_kilos_total = sold_totals['sold_kilos'] or 0
    stock_bags = max(total_bags - sold_bags_total, 0)
    stock_kilos = max(total_kilos - sold_kilos_total, 0)

    all_time_sales_total = SaleItem.objects.aggregate(total=Sum(F('unit_price') * F('quantity'), output_field=FloatField()))['total'] or 0
    fallback_bag_price = Product.objects.filter(active=True).aggregate(avg_price=Avg('sale_price'))['avg_price'] or 0
    average_value_per_bag = (all_time_sales_total / sold_bags_total) if sold_bags_total else fallback_bag_price
    sales_objective = stock_bags * average_value_per_bag

    # Ensure previous month summary is stored
    prev_month = first_day_month.month - 1 or 12
    prev_year = first_day_month.year if first_day_month.month != 1 else first_day_month.year - 1
    ms, created = MonthlySummary.objects.get_or_create(year=prev_year, month=prev_month)
    if created:
        # compute aggregates for previous month
        prev_sales_qs = SaleItem.objects.filter(sale__sale_date__year=prev_year, sale__sale_date__month=prev_month)
        prev_exp_qs = Expense.objects.filter(date__year=prev_year, date__month=prev_month)
        ms.sales_total = prev_sales_qs.aggregate(total=Sum(F('unit_price') * F('quantity'), output_field=FloatField()))['total'] or 0
        ms.expenses_total = prev_exp_qs.aggregate(total=Sum('amount'))['total'] or 0
        ms.profit_total = ms.sales_total - ms.expenses_total
        ms.bags_sold = prev_sales_qs.aggregate(total=Sum('quantity'))['total'] or 0
        ms.save()

    monthly_records = MonthlySummary.objects.all()

    context = {
        'current_month_display': current_month_display,
        'sales_total_month': format_cop(sales_total_month),
        'sales_objective': format_cop(sales_objective),
        'sales_objective_note': 'Valor del stock no vendido multiplicado por el precio promedio por bolsa.',
        'sales_objective_stock': stock_bags,
        'expenses_total_month': format_cop(expenses_total_month),
        'profit_month': format_cop(profit_month),
        'profit_month_value': profit_month,
        'margin_month': f'{margin_month:.1f} %',
        'margin_month_value': margin_month,
        'kilograms_sold_month': f'{kilograms_sold_month / 1000:.2f} kg',
        'sales_count_month': sales_count_month,
        'average_ticket': format_cop(average_ticket),
        'best_selling_product': best_selling_product.name if best_selling_product else 'N/A',
        'stock_bags': stock_bags,
        'stock_kilos': stock_kilos or 0,
        'sold_bags_total': sold_bags_total,
        'sold_kilos_total': sold_kilos_total or 0,
        'total_sales_all_time': format_cop(all_time_sales_total),
        'total_sold_bags_all_time': sold_bags_total,
        'total_expenses_all_time': format_cop(expenses_total_year),
        'total_profit_all_time': format_cop(profit_year),
        'total_bags': total_bags,
        'global_summary_note': 'Incluye ventas globales, bolsas vendidas, gastos totales y utilidad total.',
        'monthly_records': monthly_records,
        'chart_labels': json.dumps(labels, ensure_ascii=False),
        'chart_sales_data': json.dumps(sales_history),
        'chart_expenses_data': json.dumps(expenses_history),
        'chart_kg_data': json.dumps(kg_history),
        'chart_expense_labels': json.dumps([item['name'] for item in expense_distribution], ensure_ascii=False),
        'chart_expense_values': json.dumps([item['total'] for item in expense_distribution]),
    }
    return render(request, 'dashboard/home.html', context)


def export_month_csv(request, year, month):
    # Export sale items for the given month as CSV
    qs = SaleItem.objects.filter(sale__sale_date__year=year, sale__sale_date__month=month).select_related('product', 'sale')
    filename = f"ventas_{year}_{month:02d}.csv"
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    # header
    writer.writerow(['ID Venta', 'Fecha Venta', 'Producto', 'Cantidad', 'Precio Unitario', 'Total'])
    total = 0
    for item in qs:
        line_total = item.unit_price * item.quantity
        total += line_total
        writer.writerow([item.sale.id, item.sale.sale_date, item.product.name, item.quantity, item.unit_price, line_total])
    writer.writerow([])
    writer.writerow(['', '', 'Totales', '', '', total])
    return response


def export_month_xlsx(request, year, month):
    # Lazy import to avoid hard dependency at import time
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, numbers
    except ImportError:
        return HttpResponse('openpyxl is required for XLSX export', status=500)

    qs = SaleItem.objects.filter(sale__sale_date__year=year, sale__sale_date__month=month).select_related('product', 'sale')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Ventas_{year}_{month:02d}"

    headers = ['ID Venta', 'Fecha Venta', 'Producto', 'Cantidad', 'Precio Unitario', 'Total']
    header_font = Font(bold=True)
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    total = 0
    for row_idx, item in enumerate(qs, start=2):
        line_total = (item.unit_price or 0) * (item.quantity or 0)
        total += line_total
        ws.cell(row=row_idx, column=1, value=item.sale.id)
        ws.cell(row=row_idx, column=2, value=item.sale.sale_date.isoformat())
        ws.cell(row=row_idx, column=3, value=item.product.name)
        ws.cell(row=row_idx, column=4, value=item.quantity)
        ws.cell(row=row_idx, column=5, value=item.unit_price)
        ws.cell(row=row_idx, column=6, value=line_total)

    # Totals row
    total_row = qs.count() + 2
    ws.cell(row=total_row, column=5, value='Totales')
    tot_cell = ws.cell(row=total_row, column=6, value=total)
    tot_cell.font = Font(bold=True)

    # Format columns
    for col in ['E', 'F']:
        for cell in ws[col]:
            try:
                cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            except Exception:
                pass

    # Adjust column widths
    dims = {1: 10, 2: 15, 3: 30, 4: 10, 5: 15, 6: 15}
    for col, width in dims.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"ventas_{year}_{month:02d}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write(output.getvalue())
    return response
